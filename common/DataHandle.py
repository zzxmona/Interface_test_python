from openpyxl import load_workbook

# 数据请求头
class DataHandle:
    def __init__(self, testDataPath, testResultDataDir):
        self.testDataPath = testDataPath
        self.testResultDataDir = testResultDataDir
        self.wb = load_workbook(self.testDataPath)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.data = list(self.ws.values)

    def getTestData(self):
        '''
        :return: 测试数据列表
        '''
        isDoColumn = self.data[0].index('是否执行')
        testData = []
        for i in self.data[2:]:
            if i[isDoColumn] == 'y':
                tempDict = {}
                for j in range(self.ws.max_column):
                    if self.data[1][j]:
                        if not self.data[0][j]:
                            m = j
                            while not self.data[0][m]:
                                m -= 1
                            tempDict[f'{self.data[0][m]}/{self.data[1][j]}'] = i[j]
                        else:
                            tempDict[f'{self.data[0][j]}/{self.data[1][j]}'] = i[j]
                    else:
                        tempDict[self.data[0][j]] = i[j]
                testData.append(tempDict)

        return testData


    def writeData(self, list):
        for i in list:
            if len(str(i[2])) > 10000:
                fileName = f"{i[0]-2}-{i[1]}.txt"
                with open(f'{self.testResultDataDir}\{fileName}', 'w', encoding='utf-8') as f:
                    f.write(i[2])
                self.ws.cell(i[0], i[1]).value = fileName
            else:
                self.ws.cell(i[0], i[1]).value = i[2]

        self.wb.save(fr'{self.testResultDataDir}\testResult.xlsx')